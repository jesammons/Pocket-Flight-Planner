import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from datetime import datetime
import json
import xml.etree.ElementTree as ET

# Get the current date and time
current_date_and_time = datetime.now()


# Define the base URL of the website
base_url = 'https://aviationweather.gov/api/data/dataserver?requestType=retrieve&dataSource=airsigmets'
# Define excel path
airport_database = "C:\\Users\\jesam\\Documents\\GitHub\\Pocket-Flight-Planner\\airports_data.xlsx"
# Load the workbook
wb = load_workbook(airport_database)
# Get the active sheet
sheet = wb.active

# Define the search query
departing_ap = input("Enter your departing airport.: ")
# arriving_ap = input("Enter your arriving airport.:")
# Add K to airport code for AviationWeather search
query_to_code_dep = "K" + departing_ap
# convert to uppercase for excel search
uppercase_code_dep = departing_ap.upper()
# uppercase_code_arr = arriving_ap.upper()

# Find latitude and longitude from the Excel sheet
for row in sheet.iter_rows(values_only=True):
    if row[1] == uppercase_code_dep:
        departing_lat = float(row[3])  # Latitude is in the fourth column (column D)
        departing_lon = float(row[4])  # Longitude is in the fifth column (column E)
        airport_city = row[0]  # City names are in the first column (column A)
        break
else:
    print(f"Airport code '{uppercase_code_dep}' not found in the database.")
    exit()

departure_coordinates = departing_lat, departing_lon
bounding = {departing_lat + 5, departing_lon + 5, departing_lat - 5, departing_lon -5}
# arrival_coordinates = arriving_lat, arriving_lon

# Define the search parameters
search_params = {
'startTime': datetime.now(),
'endTime': datetime.now(),
'hoursBeforeNow': '1',
'format' : 'xml',
'boundingBox': '40,-90,45,-85', 
# 'q': query_to_code_dep  # Including the search query parameter if necessary
}

# Send an HTTP GET request to the API endpoint with the search parameters
response = requests.get(base_url, params=search_params)

# Check if the request was successful (status code 200)
if response.status_code == 200:
    # Process the search results or perform any desired actions
    print("Search was successful. Processing the search results here.")
    metar_data = response.content
    print("Response content:", metar_data)

    # Parse XML
    root = ET.fromstring(metar_data)
    
    airsigmet_elements = root.findall('.//AIRSIGMET')

    if airsigmet_elements:
        latest_valid_time = datetime.min
        most_relevant_airsigmet = None
    
        for airsigmet in airsigmet_elements:
            valid_time_to_str = airsigmet.find('valid_time_to').text
            valid_time_to = datetime.fromisoformat(valid_time_to_str[:-1])  # Convert to datetime object
        
            if valid_time_to > latest_valid_time:
                latest_valid_time = valid_time_to
                most_relevant_airsigmet = airsigmet
    
        if most_relevant_airsigmet:
            raw_text = most_relevant_airsigmet.find('raw_text').text
            valid_time_from = most_relevant_airsigmet.find('valid_time_from').text
            valid_time_to = most_relevant_airsigmet.find('valid_time_to').text
            altitude_min = most_relevant_airsigmet.find('altitude').get('min_ft_msl')
            altitude_max = most_relevant_airsigmet.find('altitude').get('max_ft_msl')
            hazard_type = most_relevant_airsigmet.find('hazard').get('type')
            severity = most_relevant_airsigmet.find('hazard').get('severity')
                            
        # Print the most relevant section
        print("Most relevant section based on time:")
        print(f"Tag name: AIRSIGMET")
        print(f"Raw Text:", raw_text)
        print(f"Valid Time From: ", valid_time_from)
        print(f"Valid Time To: ", valid_time_to )
        print(f"Altitude Min:", altitude_min)
        print(f"Altitude Max:", altitude_max)
        print(f"Hazard Type: ", hazard_type)
        print(f"Severity: ", severity)
        print("-" * 50)
else:
    print("Failed to retrieve data from the API.")
