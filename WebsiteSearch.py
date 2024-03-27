import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import json
from geopy.distance import geodesic
from datetime import datetime
import xml.etree.ElementTree as ET
import folium
from io import BytesIO
import base64
import numpy as np

# Get the current date and time
current_date_and_time = datetime.now()


# Define the base URL of the website
base_url = 'https://aviationweather.gov/api/data/taf'
# Define the base URL of the website
weather_url = 'https://api.openweathermap.org/data/2.5/forecast?appid=3bb35aef1eba1352e69736edcd7ab79f&cnt=1&units=metric'
# Define excel path
airport_database = "C:\\Users\\jesam\\Documents\\GitHub\\Pocket-Flight-Planner\\airports_data.xlsx"
# Load the workbook
wb = load_workbook(airport_database)
# Get the active sheet
sheet = wb.active

def openweathersearch (lats, lons) : 
    # Begin OpenWeatherAPI search

    # Define search parameters for weather forecast
    weather_params1 = {
        'lat': lats,
        'lon': lons,
        'appid': '3bb35aef1eba1352e69736edcd7ab79f',
        'mode': 'json',
        'units': 'metric',
        'lang': 'english'
    }
    # Send an HTTP GET request to the weather API endpoint
    response1 = requests.get(weather_url, params=weather_params1)
    # Check if the request was successful (status code 200)
    if response1.status_code == 200:
        # Process the weather forecast results
        print("Weather Forecast Search Results:")
        openweather_data = response1.content
        # Assuming openweather_data contains the JSON response
        response_data = json.loads(openweather_data)
        # print("Response content:", openweather_data)
        
        # Check if the 'list' key is present in the response data
        if 'list' in response_data:
            # Extract the first object from the 'list'
            first_weather_data = response_data['list'][0]

            # Check if the 'main' key is present in the first weather data
            if 'main' in first_weather_data:
                main_data = first_weather_data['main']

                # Extract temperature-related information
                temperature = main_data.get('temp')
                feels_like = main_data.get('feels_like')
                temp_min = main_data.get('temp_min')
                temp_max = main_data.get('temp_max')
                pressure = main_data.get('pressure')
                sea_level = main_data.get('sea_level')
                ground_level = main_data.get('grnd_level')
                humidity = main_data.get('humidity')
                # Extract weather description
                weather_description = ""
                if 'weather' in first_weather_data and len(first_weather_data['weather']) > 0:
                    cloud_description = first_weather_data['weather'][0].get('description')
                # Extract wind-related information
                wind = first_weather_data.get('wind')
                wind_speed = wind.get('speed') if wind else None
                wind_degrees = wind.get('deg') if wind else None
                wind_gust = wind.get('gust') if wind else None
                visibility = main_data.get ('visibility')
                # Extract city name
                county = response_data['city']['name'] if 'city' in response_data else None

                # Print the extracted information
                print("Temperature:", temperature, "degrees Celsius")
                print("Feels Like:", feels_like, "degrees Celsius")
                print("Minimum Temperature:", temp_min, "degrees Celsius")
                print("Maximum Temperature:", temp_max, "degrees Celsius")
                print("Pressure:", pressure)
                print("Sea level:", sea_level)
                print("Ground level:", ground_level)
                print("Humidity:", humidity, "%")
                print("Clouds:", cloud_description)
                print("Wind:", wind_speed, "m/s, from:" , wind_degrees, "degrees, gusting to: ", wind_gust, "m/s" )
                print("visibility:", visibility)
                print("county:", county)
                print("-" * 50)

        else:
            print("No 'list' data found in the response.")
    else:
        print("Failed to perform the weather forecast search. Status code:", response1.status_code)
    
# End OpenWeatherAPI search

        
def intermediate_lat_lon(departure_coordinates, arrival_coordinates): 
        # After obtaining departure_coordinates and arrival_coordinates
        intermediate_lats, intermediate_lons = calculate_intermediate_points(departure_coordinates, arrival_coordinates)

        print("\nLatitude, Longitude:")
        for lat, lon in zip(intermediate_lats, intermediate_lons):
            print(f"{lat:.6f}, {lon:.6f}")

        # print(f"\nDistance between {departing_ap} and {arriving_ap}: {distance_miles:.2f} miles")
        # print(f"Estimated flight time: {flight_time_hours:.2f} hours")

        # Calculate bounding box for the departing and arriving airports
        min_lon = min(departing_lon, arriving_lon) - 5
        max_lon = max(departing_lon, arriving_lon) + 5
        min_lat = min(departing_lat, arriving_lat) - 5
        max_lat = max(departing_lat, arriving_lat) + 5
        
def calculate_intermediate_points(departure_coordinates, arrival_coordinates, interval_miles=20):
    distance_miles = calculate_distance(departure_coordinates, arrival_coordinates)
    num_points = int(distance_miles / interval_miles) + 1

    lats, lons = [], []

    for i in range(num_points):
        fraction = i / (num_points - 1)
        # Assuming departure_coordinates and arrival_coordinates are tuples or lists of coordinates
        # Convert string coordinates to float
        departure_coords = (float(departure_coordinates[0]), float(departure_coordinates[1]))
        arrival_coords = (float(arrival_coordinates[0]), float(arrival_coordinates[1]))
        # Then perform arithmetic operations
        intermediate_point = (
            departure_coords[0] + fraction * (arrival_coords[0] - departure_coords[0]),
            departure_coords[1] + fraction * (arrival_coords[1] - departure_coords[1])
        )
        lats.append(intermediate_point[0])
        lons.append(intermediate_point[1])

        openweathersearch(lats[i], lons[i])

    return lats, lons
        
def calculate_distance(coord1, coord2):
    return geodesic(coord1, coord2).miles

def aviationweathersearch (arrival_coordinates, departure_coordinates) : 
    # Example using a predefined boundary for the United States
    us_boundary = '24,-125,49,-66'  # ( min_latitude,min_longitude, max_latitude, max_longitude)
    chicago_boundary = '40,-90,45,-85'
    
    # Define the base URL of the website
    base_url = 'https://aviationweather.gov/api/data/dataserver?requestType=retrieve&dataSource=airsigmets'

    buffer_distance = 5  # Adjust this value as needed
    bounding = [
        departure_coordinates[0] - buffer_distance,  # Min latitude
        departure_coordinates[1] - buffer_distance,   # Min longitude
        arrival_coordinates[0] + buffer_distance,  # Max latitude
        arrival_coordinates[1] + buffer_distance  # Max longitude
    ]
    bounding_box_str = ','.join(map(str, bounding))

    # Define the search parameters
    search_params = {
    'startTime': datetime.now(),
    'endTime': datetime.now(),
    'hoursBeforeNow': '1',
    'format' : 'xml',
    'boundingBox': bounding_box_str, #bounding box surrouding United States
    # 'q': query_to_code_dep  # Including the search query parameter if necessary
    }

    # Send an HTTP GET request to the API endpoint with the search parameters
    response = requests.get(base_url, params=search_params)

    # Check if the request was successful (status code 200)
    if response.status_code == 200:
        # Process the search results or perform any desired actions
        print("Search was successful. Processing the search results here.")
        metar_data = response.content
        # print("Response content:", metar_data)

        # Parse XML
        root = ET.fromstring(metar_data)
        
        airsigmet_elements = root.findall('.//AIRSIGMET')

        if airsigmet_elements:
            for airsigmet in airsigmet_elements:
                raw_text = airsigmet.find('raw_text').text
                valid_time_from = airsigmet.find('valid_time_from').text
                valid_time_to = airsigmet.find('valid_time_to').text
                altitude_element = airsigmet.find('altitude')
                if altitude_element is not None:
                    altitude_min = altitude_element.get('min_ft_msl')
                    altitude_max = altitude_element.get('max_ft_msl')
                hazard_type = airsigmet.find('hazard').get('type')
                severity = airsigmet.find('hazard').get('severity')
                # latitude = airsigmet.find('latitude').text
                # longitude = airsigmet.find('longitude').text
                                
            print(f"Tag name: AIRSIGMET")
            # print(f"Raw Text:", raw_text)
            print(f"Valid Time From: ", valid_time_from)
            print(f"Valid Time To: ", valid_time_to )
            print(f"Altitude Min:", altitude_min)
            print(f"Altitude Max:", altitude_max)
            print(f"Hazard Type: ", hazard_type)
            print(f"Severity: ", severity)
            # print(f"Latitude:", latitude, "Longitude:", longitude)
            print("-" * 50)
    else:
        print("Failed to retrieve data from the API.")        

def PlaneDatabaseSearch () :
    # Define excel path
    plane_database = "C:\\Users\\jesam\\Documents\\GitHub\\Pocket-Flight-Planner\\FAA_aircraft_database.xlsx"
    # Load the workbook
    wb = load_workbook(plane_database)
    # Get the active sheet
    sheet = wb.active

    plane = input("please input the plane model:" ) 

    # Find size and weight from the Excel sheet
    for row in sheet.iter_rows(values_only=True):
        if row[4] == plane: #model is in column E
            size = row[25]  # ICAO_WTC is in column Z
            weight = row[28]  # FAA_Weight is column AB
            print("Model " , plane, "is of ICAO_WTC size ", size, "and FAA_Weight classification", weight)
            break
    else:
        print(f"Model '{plane}' not found in the database.")
        exit()

def displayweathermap() :
    # Define the base URLs of the website
    weathermap_url = 'https://tile.openweathermap.org/map/precipitation_new/{z}/{x}/{y}.png?appid=3bb35aef1eba1352e69736edcd7ab79f'
    cloudmap_url = 'https://tile.openweathermap.org/map/clouds_new/{z}/{x}/{y}.png?appid=3bb35aef1eba1352e69736edcd7ab79f'
    windmap_url = 'https://tile.openweathermap.org/map/wind_new/{z}/{x}/{y}.png?appid=3bb35aef1eba1352e69736edcd7ab79f'
    pressuremap_url = 'https://tile.openweathermap.org/map/pressure_new/{z}/{x}/{y}.png?appid=3bb35aef1eba1352e69736edcd7ab79f'
    # Define the coordinates for the United States
    us_coordinates = [[69, -145], [15, -45]]

    # Create a folium map centered at (37.0902, -95.7129) with zoom level 5
    m = folium.Map(location=[37.0902, -95.7129], zoom_start=5)

    # Send an HTTP GET request to the weather API endpoints
    response1 = requests.get(weathermap_url.format(z=1, x=1, y=0))
    response2 = requests.get(cloudmap_url.format(z=1, x=1, y=1))
    response3 = requests.get(windmap_url.format(z=1, x=1, y=0))
    response4 = requests.get(pressuremap_url.format(z=1, x=1, y=0))

    # Check if the requests were successful (status code 200)
    if response1.status_code == 200 and response2.status_code == 200 and response3.status_code == 200 and response4.status_code == 200:
        # Overlay precipitation data
        precip_data = response1.content
        precip_image_url = f"data:image/png;base64,{base64.b64encode(precip_data).decode()}"
        folium.raster_layers.ImageOverlay(
            image=precip_image_url,
            bounds=us_coordinates,
            opacity=1.0,  # Adjust opacity as needed
        ).add_to(m)

        # Overlay cloud data
        cloud_data = response2.content
        cloud_image_url = f"data:image/png;base64,{base64.b64encode(cloud_data).decode()}"
        folium.raster_layers.ImageOverlay(
            image=cloud_image_url,
            bounds=us_coordinates,
            opacity=0.5,  # Adjust opacity as needed
        ).add_to(m)

        # Overlay wind data
        wind_data = response3.content
        wind_image_url = f"data:image/png;base64,{base64.b64encode(wind_data).decode()}"
        folium.raster_layers.ImageOverlay(
            image=wind_image_url,
            bounds=us_coordinates,
            opacity=0.8,  # Adjust opacity as needed
        ).add_to(m)
        
        # Overlay pressure data
        pressure_data = response4.content
        pressure_image_url = f"data:image/png;base64,{base64.b64encode(pressure_data).decode()}"
        folium.raster_layers.ImageOverlay(
            image=pressure_image_url,
            bounds=us_coordinates,
            opacity=0.3,  # Adjust opacity as needed
        ).add_to(m)

        # Save the map to an HTML file
        m.save('weather_forecast_map.html')
        # Open the HTML file in the default web browser
        import webbrowser
        webbrowser.open('weather_forecast_map.html')
    else:
        print("Failed to perform the weather forecast search. Status codes:", response1.status_code, response2.status_code)

# Main code            
# Define the search query
departing_ap = input("Enter your departing airport.: ")
arriving_ap = input("Enter your arriving airport.:")
# Add K to airport code for AviationWeather search
query_to_code_dep = "K" + departing_ap
query_to_code_arr = "K" + arriving_ap
# convert to lowercase for excel search
uppercase_code_dep = departing_ap.upper()
uppercase_code_arr = arriving_ap.upper()
# Find latitude and longitude from the Excel sheet
for row in sheet.iter_rows(values_only=True):
    if row[1] == uppercase_code_dep:
        departing_lat = float(row[3])  # Latitude is in the fourth column (column D)
        departing_lon = float(row[4])  # Longitude is in the fifth column (column E)
        airport_city = row[0]  # City names are in the first column (column A)
        break
else:
    print(f"Airport code '{uppercase_code_dep}' not found in the database.")
    exit()
# Find latitude and longitude from the Excel sheet
for row in sheet.iter_rows(values_only=True):
    if row[1] == uppercase_code_arr:
        arriving_lat = float(row[3])  # Latitude is in the fourth column (column D)
        arriving_lon = float(row[4])  # Longitude is in the fifth column (column E)
        airport_city = row[0]  # City names are in the first column (column A)
        break
else:
    print(f"Airport code '{uppercase_code_arr}' not found in the database.")
    exit()

# Call the functions
departure_coordinates = (departing_lat, departing_lon)
arrival_coordinates = (arriving_lat, arriving_lon)
calculate_intermediate_points(departure_coordinates, arrival_coordinates)
aviationweathersearch(arrival_coordinates, departure_coordinates)
PlaneDatabaseSearch()
displayweathermap()